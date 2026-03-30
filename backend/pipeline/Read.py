import pandas as pd
import math
import subprocess
import os
import io
import openpyxl
import msoffcrypto

HEADER_ROW = 3

# =========================
# MASTER RECIPE CONFIG
# =========================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

MASTER_RECIPES = {
    "KETTLEANDBLENDTECH": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "KETTLEANDBLENDTECH.xlsx"),
        "password": "NPD04",
    },
    "KETTLE1SOUP": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "SOUP.xlsx"),
        "password": "NPD06",
    },
    "KETTLE2SAUCE": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "SAUCE.xlsx"),
        "password": "NPD05",
    },
    "VAPASTRY": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "PASTRY.xlsx"),
        "password": "NPD11",
    },
    "MARINATION": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "MARINATION.xlsx"),
        "password": "NPD10",
    },
    "WOK": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "WOK.xlsx"),
        "password": "NPD12",
    },
    "RICE": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "RICE.xlsx"),
        "password": "NPD09",
    },
    "CQC": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "CQC.xlsx"),
        "password": "NPD01",
    },
}

# Mature cheese WIP codes that need adjustment
MATURE_CHEESE_CODES = [
    "WFMCS10000",   # Coles Finest Mature Cheese Sauce
]

# =========================
# MATURE CHEESE HELPERS
# =========================

def decrypt_excel(path: str, password: str) -> io.BytesIO:
    """Decrypt password-protected Excel file to memory."""
    buffer = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(buffer)
    buffer.seek(0)
    return buffer

def load_master(key: str):
    """Load master recipe from memory."""
    cfg = MASTER_RECIPES[key]
    decrypted = decrypt_excel(cfg["file"], cfg["password"])
    return openpyxl.load_workbook(decrypted, data_only=True)

def build_recipe_map():
    """Build mapping of WIP codes to master sheets."""
    mapping = {}
    for key in MASTER_RECIPES:
        wb = load_master(key)
        ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.worksheets[0]
        
        for r in range(2, ws.max_row + 1):
            wip = ws.cell(r, 2).value
            ref = ws.cell(r, 3)
            if not wip:
                continue
            
            if ref.hyperlink and ref.hyperlink.location:
                sheet = ref.hyperlink.location.split("!")[0].strip("'")
            else:
                sheet = ref.value
            
            if sheet:
                mapping[str(wip).strip()] = {"master": key, "sheet": sheet}
        
        wb.close()
    return mapping

def find_total_in_and_finished_yield(ws):
    total_in_value = None
    finished_yield_value = None
    
    for r in range(1, min(ws.max_row + 1, 50)):
        desc = ws.cell(r, 3).value  # Column C - Ingredient Description (was 2, now 3)
        control = ws.cell(r, 4).value  # Column D - Control
        
        if desc:
            desc_lower = str(desc).lower().strip()
            
            # Check for "Total in"
            if "total" in desc_lower and "in" in desc_lower:
                
                if control is not None:
                    try:
                        total_in_value = float(control)
                    except:
                        pass
            
            # Check for "Finished yield"
            if "finished" in desc_lower and "yield" in desc_lower:
                
                if control is not None:
                    try:
                        finished_yield_value = float(control)
                    except:
                        pass
    
    
    return total_in_value, finished_yield_value

def calculate_mature_cheese_batch_size(item_code, target_batch, recipe_map):
    
    if item_code not in recipe_map:
        return target_batch
    
    recipe_info = recipe_map[item_code]
    master_key = recipe_info["master"]
    sheet_name = recipe_info["sheet"]
    
    try:
        wb = load_master(master_key)
        ws = wb[sheet_name]
        
        total_in, finished_yield = find_total_in_and_finished_yield(ws)
        
        wb.close()
        
        if total_in and finished_yield and total_in > 0:
            # Formula: Adjusted Batch = Target × (Finished Yield / Total In)
            adjusted = target_batch * (finished_yield / total_in)
            return adjusted
        else:
            return target_batch
    
    except Exception as e:
        return target_batch

# =========================
# READ EXCEL FILE
# =========================

if HEADER_ROW == 0:
    df = pd.read_excel('production.xlsx')
else:
    df = pd.read_excel('production.xlsx', header=HEADER_ROW)

# =========================
# STANDARDISE COLUMN NAMES
# =========================

column_mapping = {}

for col in df.columns:
    col_str = str(col).strip().strip('"')
    col_normalized = col_str.lower().replace(' ', '').replace('_', '').replace('\n', '')

    if col_normalized == 'line':
        column_mapping[col] = 'Line'
    elif col_normalized == 'product':
        column_mapping[col] = 'Product'
    elif col_normalized == 'itemcode':
        column_mapping[col] = 'Item code'
    elif col_normalized in ('qty', 'quantity'):
        column_mapping[col] = 'Qty'
    elif col_normalized == 'uom':
        column_mapping[col] = 'UOM'
    elif col_normalized == 'type':
        column_mapping[col] = 'Type'
    elif col_normalized in ('comment', 'comments'):
        column_mapping[col] = 'Comments'

df = df.rename(columns=column_mapping)

required = ['Line', 'Product', 'Item code', 'Qty', 'UOM', 'Type', 'Comments']
missing = [c for c in required if c not in df.columns]
if missing:
    print(f"Error: Missing columns {missing}")
    exit()

# =========================
# CLEAN QTY COLUMN
# =========================

df = df[df['Qty'].notna()]
df = df[df['Qty'] != '']
df = df[df['Qty'] != 0]

df['Qty'] = pd.to_numeric(df['Qty'], errors='coerce')
df = df[df['Qty'].notna()]

df = df[required].copy()

print(f"\nSuccessfully loaded {len(df)} recipes from production.xlsx")

# =========================
# SAVE LINE-SPECIFIC FILES
# =========================

cqc_rice_df = df[
    df['Line'].str.lower().str.startswith('cqc') |
    (df['Line'].str.lower() == 'rice kettle')
]

oven_wok_df = df[
    (df['Line'].str.lower() == 'dd oven') |
    (df['Line'].str.lower() == 'wok')
]

if not cqc_rice_df.empty:
    cqc_rice_df.to_excel('CQC_RiceKettle.xlsx', index=False)

if not oven_wok_df.empty:
    oven_wok_df.to_excel('DDOven_Wok.xlsx', index=False)

# =========================
# BATCHING CRITERIA
# =========================

criteria = [
    'blendtech',
    'kettle 1 soup',
    'kettle 2 direct fill',
    'kettle 3 direct fill',
    'kettle 4 kapcold'
]

# EB SAUCE recipes (use bechamel batching rules)
eb_sauce_recipes = [
    'WBLSCCP10000',
]

# Exception recipes (1000 kg max batch size)
exception_recipes = [
    'WVLSCI10000',
    'WCNSOUP1000',
    'WCKMINE1000',
    'WVMSOUP1000',
    'WCVLSS10000',
]

# Special recipes with 1900 kg max batch size
special_1900_recipes = [
    'WMTSCP10000', #EM SAUCE
]

df['Line_lower'] = df['Line'].str.lower()
filtered_df = df[df['Line_lower'].isin(criteria)].copy()
filtered_df.drop(columns='Line_lower', inplace=True)

print(f"\nFound {len(filtered_df)} recipes to be batched")

# Build recipe map for mature cheese adjustment

recipe_map = build_recipe_map()
print(f"Mapped {len(recipe_map)} recipes")

# =========================
# HELPER FUNCTIONS
# =========================

def merge_batches(batches):

    if not batches:
        return ""
    
    result = []
    current = batches[0]
    count = 1

    for b in batches[1:]:
        if b == current:
            count += 1
        else:
            if count > 1:
                result.append(f"{current:.0f}*{count}")
            else:
                result.append(f"{current:.0f}")
            current = b
            count = 1

    if count > 1:
        result.append(f"{current:.0f}*{count}")
    else:
        result.append(f"{current:.0f}")
    
    return "/".join(result)


def bechamel_batches(qty):
    """Bechamel batching logic with 1880kg primary batch"""
    PRIMARY = 1880
    MIN_BATCH = 800
    
    if qty <= 2050:
        return f"{qty:.0f}"
    
    batches = []
    remaining = qty
    
    while remaining > 3760:
        batches.append(PRIMARY)
        remaining -= PRIMARY
    
    if remaining <= 2050:
        batches.append(remaining)
    else:
        remainder = remaining - PRIMARY
        if remainder >= 1000:
            batches.extend([PRIMARY, remainder])
        else:
            half = remaining / 2
            if half < MIN_BATCH:
                return f"{qty:.0f}"
            batches.extend([half, half])
    
    for i, b in enumerate(batches):
        if b < MIN_BATCH and len(batches) > 1:
            combined = batches[-2] + batches[-1]
            batches = batches[:-2]
            half = combined / 2
            batches.extend([half, half])
            break
    
    return merge_batches(batches)


def exceptional_recipe_batches(qty, max_batch=1000):
    """Exception recipes with 1000kg max batch"""
    if qty <= max_batch:
        return f"{qty:.0f}"

    n = 2
    while True:
        batch_size = qty / n
        if batch_size <= max_batch:
            if n > 1:
                return f"{batch_size:.0f}*{n}"
            else:
                return f"{batch_size:.0f}"
        n += 1


def special_1900_recipe_batches(qty):
    """Special recipes with 1900kg max batch"""
    max_batch = 1900
    batches = []
    
    if qty <= max_batch:
        return f"{qty:.0f}"
    
    if qty <= 3800:
        half = qty / 2
        return f"{half:.0f}*2"
    
    remaining = qty
    while remaining > 3800:
        batches.append(1900)
        remaining -= 1900
    
    if remaining > 1900:
        half = remaining / 2
        batches.extend([half, half])
    else:
        batches.append(remaining)
    
    return merge_batches(batches)


def normal_recipe_batches(qty):
    """Normal recipes with 2000kg max batch"""
    max_batch = 2000
    batches = []
    
    if qty <= max_batch:
        return f"{qty:.0f}"
    
    if qty <= 4000:
        half = qty / 2
        return f"{half:.0f}*2"
    
    remaining = qty
    while remaining > 4000:
        batches.append(2000)
        remaining -= 2000
    
    if remaining > 2000:
        half = remaining / 2
        batches.extend([half, half])
    else:
        batches.append(remaining)
    
    return merge_batches(batches)


def calculate_batch_size(qty, item_code):
    """Main batching logic router"""
    if item_code in eb_sauce_recipes:
        return bechamel_batches(qty)
    elif item_code in exception_recipes:
        return exceptional_recipe_batches(qty)
    elif item_code in special_1900_recipes:
        return special_1900_recipe_batches(qty)
    else:
        return normal_recipe_batches(qty)

# =========================
# APPLY BATCHING
# =========================

if not filtered_df.empty:
    print("\nCalculating batch sizes...")
    filtered_df['batch size'] = filtered_df.apply(
        lambda r: calculate_batch_size(r['Qty'], r['Item code']),
        axis=1
    )
    
    # Adjust batch sizes for mature cheese recipes

    
    def adjust_for_mature_cheese(row):
        item_code = row['Item code']
        batch_string = row['batch size']
        
        if item_code not in MATURE_CHEESE_CODES:
            return batch_string
        
        # Parse the batch string to get individual batches
        batches = []
        groups = str(batch_string).split('/')
        
        for group in groups:
            group = group.strip()
            if '*' in group:
                parts = group.split('*')
                size = float(parts[0])
                count = int(parts[1])
                for _ in range(count):
                    batches.append(size)
            else:
                batches.append(float(group))
        
        # Adjust each batch
        adjusted_batches = []
        for batch in batches:
            adjusted = calculate_mature_cheese_batch_size(item_code, batch, recipe_map)
            adjusted_batches.append(adjusted)
        
        # Merge back into string format
        return merge_batches(adjusted_batches)
    
    filtered_df['batch size'] = filtered_df.apply(adjust_for_mature_cheese, axis=1)

    output_cols = ['Line', 'Product', 'Item code', 'Qty', 'batch size', 'UOM', 'Type', 'Comments']
    filtered_df[output_cols].to_excel('batching_recipes.xlsx', index=False)

    print("\n✓ Batching results saved to batching_recipes.xlsx")

print("\n✓ Program finished successfully!")