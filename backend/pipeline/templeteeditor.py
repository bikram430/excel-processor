
import os
from datetime import datetime
from typing import List, Dict
from dataclasses import dataclass
import openpyxl

# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCTION_PLAN = os.path.join(SCRIPT_DIR, "production.xlsx")
CQC_TEMPLATE = os.path.join(SCRIPT_DIR, "CQC Production Template.xlsx")
KAPCOLD_TEMPLATE = os.path.join(SCRIPT_DIR, "Capkold Production Template.xlsx")
OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "production_output")

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Production plan structure
PROD_DATA_START = 5
PROD_COLS = {
    "LINE": 1,          # Column A
    "PRODUCT": 2,       # Column B
    "ITEM_CODE": 3,     # Column C
    "QTY": 8,           # Column H
    "UOM": 9,           # Column I
}

# CQC keywords
CQC_KEYWORDS = ["CQC 1", "CQC 2", "CQC1", "CQC2"]
RICE_KETTLE_KEYWORDS = ["RICE KETTLE", "RICE-KETTLE", "RICEKETTLE"]

# KAPCOLD keywords
KAPCOLD_GROUP1 = ["BLENDTECH", "KETTLE 4 KAPCOLD"]
KAPCOLD_GROUP2 = ["DD OVEN", "WOK"]

# CQC Template columns
CQC_COLS = {
    "LINE": 2,          # Column B
    "PRODUCT": 3,       # Column C
    "ITEM_CODE": 4,     # Column D
    "QTY": 5,           # Column E
    "UOM": 6,           # Column F
}

# KAPCOLD Template columns
KAPCOLD_COLS = {
    "LINE": 1,          # Column A
    "PRODUCT": 2,       # Column B
    "ITEM_CODE": 3,     # Column C
    "QTY": 4,           # Column D
    "UOM": 5,           # Column E
}

# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class ProductItem:
    product: str
    item_code: str
    qty: float
    uom: str
    line: str

# =============================================================================
# READERS
# =============================================================================

def read_cqc_items() -> List[ProductItem]:

    try:
        wb = openpyxl.load_workbook(PRODUCTION_PLAN, data_only=True)
        ws = wb.active
        items = []
        
        for row in range(PROD_DATA_START, ws.max_row + 1):
            line_value = ws.cell(row, PROD_COLS["LINE"]).value
            if not line_value:
                continue
            
            line_str = str(line_value).strip().upper()
            is_cqc = any(kw.upper() in line_str for kw in CQC_KEYWORDS)
            is_rice = any(kw.upper() in line_str for kw in RICE_KETTLE_KEYWORDS)
            
            if not is_cqc and not is_rice:
                continue
            
            product = ws.cell(row, PROD_COLS["PRODUCT"]).value
            item_code = ws.cell(row, PROD_COLS["ITEM_CODE"]).value
            qty = ws.cell(row, PROD_COLS["QTY"]).value
            uom = ws.cell(row, PROD_COLS["UOM"]).value
            
            if not item_code or not qty:
                continue
            
            try:
                items.append(ProductItem(
                    product=str(product).strip() if product else "",
                    item_code=str(item_code).strip(),
                    qty=float(qty),
                    uom=str(uom).strip() if uom else "KG",
                    line="CQC" if is_cqc else "RICE_KETTLE"
                ))
            except:
                continue
        
        wb.close()
        return items
    except Exception as e:
        print(f"ERROR reading CQC items: {e}")
        return []

def read_kapcold_items() -> Dict[str, List[ProductItem]]:

    try:
        wb = openpyxl.load_workbook(PRODUCTION_PLAN, data_only=True)
        ws = wb.active
        groups = {"GROUP1": [], "GROUP2": [], "RICE_KETTLE": []}
        
        for row in range(PROD_DATA_START, ws.max_row + 1):
            line_value = ws.cell(row, PROD_COLS["LINE"]).value
            if not line_value:
                continue
            
            line_str = str(line_value).strip().upper()
            group = None
            
            for kw in KAPCOLD_GROUP1:
                if kw.upper() in line_str:
                    group = "GROUP1"
                    break
            
            if not group:
                for kw in KAPCOLD_GROUP2:
                    if kw.upper() in line_str:
                        group = "GROUP2"
                        break
            
            if not group:
                for kw in RICE_KETTLE_KEYWORDS:
                    if kw.upper() in line_str:
                        group = "RICE_KETTLE"
                        break
            
            if not group:
                continue
            
            product = ws.cell(row, PROD_COLS["PRODUCT"]).value
            item_code = ws.cell(row, PROD_COLS["ITEM_CODE"]).value
            qty = ws.cell(row, PROD_COLS["QTY"]).value
            uom = ws.cell(row, PROD_COLS["UOM"]).value
            
            if not item_code or not qty:
                continue
            
            try:
                groups[group].append(ProductItem(
                    product=str(product).strip() if product else "",
                    item_code=str(item_code).strip(),
                    qty=float(qty),
                    uom=str(uom).strip() if uom else "KG",
                    line=line_str
                ))
            except:
                continue
        
        wb.close()
        return groups
    except Exception as e:
        print(f"ERROR reading KAPCOLD items: {e}")
        return {}

# =============================================================================
# FILLERS
# =============================================================================

def fill_cqc_template(items: List[ProductItem]) -> str:

    try:
        wb = openpyxl.load_workbook(CQC_TEMPLATE)
        ws = wb.active
        
        # Unhide all rows
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].hidden = False
        
        # Update date
        date_str = datetime.now().strftime('%d/%m/%Y')
        for cell_ref in ['I2', 'J2', 'K2']:
            try:
                ws[cell_ref] = date_str
                break
            except:
                pass
        
        # Clear data
        for row in range(6, ws.max_row + 1):
            for col in [2, 3, 4, 5, 6]:
                ws.cell(row, col).value = None
        
        # Separate items
        cqc_items = [item for item in items if item.line == "CQC"]
        rice_items = [item for item in items if item.line == "RICE_KETTLE"]
        
        # Write CQC
        current_row = 6
        line_number = 1
        for item in cqc_items:
            ws.cell(current_row, CQC_COLS["LINE"]).value = line_number
            ws.cell(current_row, CQC_COLS["PRODUCT"]).value = item.product
            ws.cell(current_row, CQC_COLS["ITEM_CODE"]).value = item.item_code
            ws.cell(current_row, CQC_COLS["QTY"]).value = item.qty
            ws.cell(current_row, CQC_COLS["UOM"]).value = item.uom
            line_number += 1
            current_row += 1
        
        # Gap
        current_row += 2
        
        # Write RICE KETTLE
        line_number = 1
        for item in rice_items:
            ws.cell(current_row, CQC_COLS["LINE"]).value = line_number
            ws.cell(current_row, CQC_COLS["PRODUCT"]).value = item.product
            ws.cell(current_row, CQC_COLS["ITEM_CODE"]).value = item.item_code
            ws.cell(current_row, CQC_COLS["QTY"]).value = item.qty
            ws.cell(current_row, CQC_COLS["UOM"]).value = item.uom
            line_number += 1
            current_row += 1
        
        # Hide unused rows
        for row in range(current_row, ws.max_row + 1):
            ws.row_dimensions[row].hidden = True
        
        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(OUTPUT_FOLDER, f"CQC_{timestamp}.xlsx")
        wb.save(output_file)
        wb.close()
        
        return output_file
    except Exception as e:
        print(f"ERROR filling CQC template: {e}")
        raise

def fill_kapcold_template(groups: Dict[str, List[ProductItem]]) -> str:

    try:
        wb = openpyxl.load_workbook(KAPCOLD_TEMPLATE)
        ws = wb.active
        
        # Unhide all rows
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].hidden = False
        
        # Update date
        date_str = datetime.now().strftime('%d/%m/%Y')
        for cell_ref in ['I2', 'J2', 'K2']:
            try:
                ws[cell_ref] = date_str
                break
            except:
                pass
        
        # Clear data
        for row in range(6, ws.max_row + 1):
            for col in [1, 2, 3, 4, 5]:
                ws.cell(row, col).value = None
        
        current_row = 6
        
        # GROUP 1
        line_number = 1
        for item in groups['GROUP1']:
            ws.cell(current_row, KAPCOLD_COLS["LINE"]).value = line_number
            ws.cell(current_row, KAPCOLD_COLS["PRODUCT"]).value = item.product
            ws.cell(current_row, KAPCOLD_COLS["ITEM_CODE"]).value = item.item_code
            ws.cell(current_row, KAPCOLD_COLS["QTY"]).value = item.qty
            ws.cell(current_row, KAPCOLD_COLS["UOM"]).value = item.uom
            line_number += 1
            current_row += 1
        
        current_row += 3  # Gap
        
        # GROUP 2
        line_number = 1
        for item in groups['GROUP2']:
            ws.cell(current_row, KAPCOLD_COLS["LINE"]).value = line_number
            ws.cell(current_row, KAPCOLD_COLS["PRODUCT"]).value = item.product
            ws.cell(current_row, KAPCOLD_COLS["ITEM_CODE"]).value = item.item_code
            ws.cell(current_row, KAPCOLD_COLS["QTY"]).value = item.qty
            ws.cell(current_row, KAPCOLD_COLS["UOM"]).value = item.uom
            line_number += 1
            current_row += 1
        
        current_row += 3  # Gap
        
        # RICE KETTLE
        line_number = 1
        for item in groups['RICE_KETTLE']:
            ws.cell(current_row, KAPCOLD_COLS["LINE"]).value = line_number
            ws.cell(current_row, KAPCOLD_COLS["PRODUCT"]).value = item.product
            ws.cell(current_row, KAPCOLD_COLS["ITEM_CODE"]).value = item.item_code
            ws.cell(current_row, KAPCOLD_COLS["QTY"]).value = item.qty
            ws.cell(current_row, KAPCOLD_COLS["UOM"]).value = item.uom
            line_number += 1
            current_row += 1
        
        # Hide unused rows
        for row in range(current_row, ws.max_row + 1):
            ws.row_dimensions[row].hidden = True
        
        # Save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(OUTPUT_FOLDER, f"KAPCOLD_{timestamp}.xlsx")
        wb.save(output_file)
        wb.close()
        
        return output_file
    except Exception as e:
        print(f"ERROR filling KAPCOLD template: {e}")
        raise

# =============================================================================
# MAIN
# =============================================================================

def main():

    
    try:
        # CQC
        print("\nProcessing CQC...")
        cqc_items = read_cqc_items()
        if cqc_items:
            cqc_file = fill_cqc_template(cqc_items)
            print(f"✓ CQC: {os.path.basename(cqc_file)}")
        else:
            print("✗ CQC: No items found")
        
        # KAPCOLD
        print("\nProcessing KAPCOLD...")
        kapcold_groups = read_kapcold_items()
        if any(kapcold_groups.values()):
            kapcold_file = fill_kapcold_template(kapcold_groups)
            print(f"✓ KAPCOLD: {os.path.basename(kapcold_file)}")
        else:
            print("✗ KAPCOLD: No items found")
        
        print(f"\n DONE! Files saved to: {OUTPUT_FOLDER}")
    
    except Exception as e:
        print(f"\n FAILED: {e}")
        import traceback
        traceback.print_exc()
    
    print("=" * 60)

if __name__ == "__main__":
    main()
