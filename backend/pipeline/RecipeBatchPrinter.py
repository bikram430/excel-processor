import io
import os
import re
from copy import copy
from typing import Dict, List, Tuple

import openpyxl
import msoffcrypto

# =============================================================================
# CONFIG
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_PRINT_FOLDER = os.path.join(SCRIPT_DIR, "temp_print")
BATCHING_FILE = os.path.join(SCRIPT_DIR, "batching_recipes.xlsx")

os.makedirs(TEMP_PRINT_FOLDER, exist_ok=True)

# MASTER RECIPES CONFIG
MASTER_RECIPES: Dict[str, Dict[str, str]] = {
    "KETTLEANDBLENDTECH": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "KETTLEANDBLENDTECH.xlsx"),
        "password": "NPD04",
    },
    "KETTLE1SOUP": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "SOUP.xlsx"),
        "password": "NPD06",
    },
    "KETTLE2SAUCE": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "SAUCE.xlsx"),
        "password": "NPD05",
    },
    "VAPASTRY": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "PASTRY.xlsx"),
        "password": "NPD11",
    },
    "MARINATION": {
        "file": os.path.join(SCRIPT_DIR, "Recipes", "MARINATION.xlsx"),
        "password": "NPD07",
    },
}

HEADER_ROW = 6  # Row where headers exist

# MARINATED MEAT CODES - these require marination recipe expansion
MARINATED_CODES = {
    "WBMMCP10000": "Beef Marinated",
    "WSNCCA10000": "Chicken Marinated (Singapore)",
    "WMDTCCD1000": "Marinated Diced Tandoori Chicken",
    "WRMCTD10000": "Roasted Moroccan Chicken Thigh 20mm Dices",
    "WLMMCP10000": "Lamb Marinade",
    "WRCTM10000": "Raw Chicken Tikka Marinated",
}

# =============================================================================
# HELPERS
# =============================================================================

def normalize_text(val) -> str:
    """Normalize text for comparison."""
    return str(val).strip().lower() if val else ""


def parse_batch_size(batch_string) -> List[float]:
    """Parse batch size string into list of individual batch sizes."""
    batches = []
    
    # Convert to string and split by "/" to get different batch groups
    groups = str(batch_string).split('/')
    
    for group in groups:
        group = group.strip()
        
        # Check if this group has "*" (multiplication)
        if '*' in group:
            parts = group.split('*')
            size = float(parts[0])
            count = int(parts[1])
            
            # Add 'count' number of batches of this size
            for _ in range(count):
                batches.append(size)
        else:
            # Single batch, no multiplication
            batches.append(float(group))
    
    return batches


# =============================================================================
# DECRYPTION
# =============================================================================

def decrypt_excel(path: str, password: str) -> io.BytesIO:
    """Decrypt password-protected Excel file to memory buffer."""
    buffer = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(buffer)
    buffer.seek(0)
    return buffer


def load_master(key: str, *, data_only: bool):
    """Load master recipe from memory. Original file never touched."""
    cfg = MASTER_RECIPES[key]
    decrypted = decrypt_excel(cfg["file"], cfg["password"])
    return openpyxl.load_workbook(decrypted, data_only=data_only)


# =============================================================================
# BATCH FILE READER
# =============================================================================

def read_batching_file() -> List[Dict]:
    """Read batching_recipes.xlsx to get items and batch sizes."""
    wb = openpyxl.load_workbook(BATCHING_FILE, data_only=True)
    ws = wb.active

    # Find column headers
    headers = {
        normalize_text(ws.cell(1, c).value).replace(" ", ""): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }

    def find(names):
        for n in names:
            k = normalize_text(n).replace(" ", "")
            if k in headers:
                return headers[k]
        return None

    item_col = find(["itemcode", "wipcode"])
    batch_col = find(["batchsize"])

    if not item_col or not batch_col:
        raise RuntimeError("Missing 'Item Code' or 'Batch Size' columns in batching file")

    entries = []
    for r in range(2, ws.max_row + 1):
        item = ws.cell(r, item_col).value
        if item:
            entries.append({
                "item_code": str(item).strip(),
                "batch_field": ws.cell(r, batch_col).value,
            })

    wb.close()
    return entries


# =============================================================================
# MASTER MAP BUILDER
# =============================================================================

def build_master_map() -> Dict[str, Dict[str, str]]:
    """Map WIP codes to master recipe file and sheet."""
    mapping = {}
    for key in MASTER_RECIPES:
        cfg = MASTER_RECIPES[key]
        if not os.path.exists(cfg["file"]):
            print(f"\n⚠ Skipping {key}: File not found")
            continue
            
        print(f"\n→ Reading master: {key}")
        wb = load_master(key, data_only=False)
        
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
        else:
            ws = wb.worksheets[0]
            print(f"  ⚠ No Summary sheet, using: {ws.title}")
        
        found_count = 0
        for r in range(2, min(ws.max_row + 1, 300)):  # Scan up to row 300
            wip = ws.cell(r, 2).value  # Column B
            ref = ws.cell(r, 3).value  # Column C
            
            sheet = None
            # Check for hyperlink first
            ref_cell = ws.cell(r, 3)
            if ref_cell.hyperlink and ref_cell.hyperlink.location:
                sheet = ref_cell.hyperlink.location.split("!")[0].strip("'")
            elif ref:
                sheet = str(ref).strip()
            
            if sheet and wip:
                wip_code = str(wip).strip()
                mapping[wip_code] = {"master": key, "sheet": sheet}
                found_count += 1
        
        print(f"  ✓ Found {found_count} recipes")
        wb.close()
    return mapping


# =============================================================================
# FIND RECIPE STRUCTURE
# =============================================================================

def find_recipe_structure(ws):
    """Find column positions and key rows in recipe sheet."""
    
    # Find columns by scanning header row
    mix_col = None
    ingredient_code_col = None
    ingredient_desc_col = None
    control_col = None
    percent_col = None
    new_batch_col = None
    
    for col in range(1, 20):
        header = ws.cell(HEADER_ROW, col).value
        if not header:
            continue
        text = normalize_text(header)
        
        if text == "mix":
            mix_col = col
        elif "ingredient" in text and "code" in text:
            ingredient_code_col = col
        elif "ingredient" in text and ("description" in text or "desc" in text):
            ingredient_desc_col = col
        elif text == "control":
            control_col = col
        elif text == "%":
            percent_col = col
        elif "new" in text and "batch" in text:
            new_batch_col = col
    
    if not all([control_col, percent_col, new_batch_col]):
        raise RuntimeError(f"Missing required columns. Found: Control={control_col}, %={percent_col}, New Batch={new_batch_col}")
    
    # Find ingredient row range
    ingredient_start = HEADER_ROW + 1
    ingredient_end = None
    
    # Look for "Total in" to mark end of ingredients
    for r in range(ingredient_start, min(ingredient_start + 50, ws.max_row + 1)):
        cell_val = ws.cell(r, ingredient_desc_col).value if ingredient_desc_col else ws.cell(r, control_col).value
        if cell_val:
            text = normalize_text(str(cell_val))
            if "total" in text and "in" in text:
                ingredient_end = r - 1
                break
    
    if not ingredient_end or ingredient_end < ingredient_start:
        ingredient_end = ingredient_start + 15  # Default to 15 ingredients
    
    # Find summary rows
    total_in_row = None
    finished_yield_row = None
    yield_loss_row = None
    
    search_start = ingredient_end + 1
    for r in range(search_start, min(search_start + 10, ws.max_row + 1)):
        desc = ws.cell(r, ingredient_desc_col).value if ingredient_desc_col else None
        if desc:
            text = normalize_text(desc)
            if "total" in text and "in" in text:
                total_in_row = r
            elif "finished" in text and "yield" in text:
                finished_yield_row = r
            elif "yield" in text and "loss" in text:
                yield_loss_row = r
    
    # Find batch size cell (in header area, rows 1-5)
    batch_size_row = None
    batch_size_col = None
    for r in range(1, HEADER_ROW):
        for c in range(1, 10):
            cell_val = ws.cell(r, c).value
            if cell_val and "batch" in normalize_text(str(cell_val)) and "size" in normalize_text(str(cell_val)):
                batch_size_row = r
                batch_size_col = c + 1  # Value is in next column
                break
        if batch_size_row:
            break
    
    return {
        "mix_col": mix_col,
        "ingredient_code_col": ingredient_code_col,
        "ingredient_desc_col": ingredient_desc_col,
        "control_col": control_col,
        "percent_col": percent_col,
        "new_batch_col": new_batch_col,
        "ingredient_start": ingredient_start,
        "ingredient_end": ingredient_end,
        "total_in_row": total_in_row,
        "finished_yield_row": finished_yield_row,
        "yield_loss_row": yield_loss_row,
        "batch_size_row": batch_size_row,
        "batch_size_col": batch_size_col,
    }


# =============================================================================
# COPY PRINT SETTINGS
# =============================================================================

def copy_print_settings(ws_from, ws_to):
    """Copy all print and page settings from source to target worksheet."""
    
    try:
        # Copy page setup
        if ws_from.page_setup:
            ws_to.page_setup.orientation = ws_from.page_setup.orientation
            ws_to.page_setup.paperSize = ws_from.page_setup.paperSize
            
            # FORCE FIT TO PAGE
            ws_to.page_setup.fitToPage = True
            ws_to.page_setup.fitToWidth = 1
            ws_to.page_setup.fitToHeight = 0
            ws_to.page_setup.scale = None
            
        # Copy page margins
        if ws_from.page_margins:
            ws_to.page_margins.left = ws_from.page_margins.left
            ws_to.page_margins.right = ws_from.page_margins.right
            ws_to.page_margins.top = ws_from.page_margins.top
            ws_to.page_margins.bottom = ws_from.page_margins.bottom
            ws_to.page_margins.header = ws_from.page_margins.header
            ws_to.page_margins.footer = ws_from.page_margins.footer
        
        # Copy print options
        if ws_from.print_options:
            ws_to.print_options.horizontalCentered = ws_from.print_options.horizontalCentered
            ws_to.print_options.verticalCentered = ws_from.print_options.verticalCentered
            ws_to.print_options.headings = ws_from.print_options.headings
            ws_to.print_options.gridLines = ws_from.print_options.gridLines
        
        # Copy print area
        if ws_from.print_area:
            ws_to.print_area = ws_from.print_area
        
        # Copy print titles
        if hasattr(ws_from, 'print_title_rows') and ws_from.print_title_rows:
            ws_to.print_title_rows = ws_from.print_title_rows
        if hasattr(ws_from, 'print_title_cols') and ws_from.print_title_cols:
            ws_to.print_title_cols = ws_from.print_title_cols
        
    except Exception as e:
        print(f"  Warning: Could not copy all print settings: {str(e)}")


# =============================================================================
# EXTRACT MARINATED CODES FROM RECIPE
# =============================================================================

def extract_marinated_codes_from_recipe(ws_value, structure, new_batch_size: float) -> List[Tuple[str, float]]:
    """
    Scan recipe for marinated meat codes and calculate their quantities.
    Uses the formula: new_qty = control_value * (new_batch_size / finished_yield)
    Returns list of (marinated_code, quantity) tuples.
    """
    marinated_items = []
    
    code_col = structure["ingredient_code_col"]
    control_col = structure["control_col"]
    
    if not code_col or not control_col:
        return marinated_items
    
    # Get finished yield from the original recipe
    finished_yield = None
    if structure["finished_yield_row"] and control_col:
        finished_yield_val = ws_value.cell(structure["finished_yield_row"], control_col).value
        if finished_yield_val:
            try:
                finished_yield = float(finished_yield_val)
            except (ValueError, TypeError):
                pass
    
    if not finished_yield or finished_yield <= 0:
        print(f"    ⚠ Could not find finished yield, skipping marination detection")
        return marinated_items
    
    # Calculate scale ratio
    scale_ratio = new_batch_size / finished_yield
    
    for r in range(structure["ingredient_start"], structure["ingredient_end"] + 1):
        code_val = ws_value.cell(r, code_col).value
        if not code_val:
            continue
        
        code = str(code_val).strip().upper()
        
        # Check if this is a marinated code
        if code in MARINATED_CODES:
            # Get control value (original quantity)
            control_val = ws_value.cell(r, control_col).value
            if control_val:
                try:
                    control_qty = float(control_val)
                    if control_qty > 0:
                        # Calculate new quantity using the formula
                        new_qty = control_qty * scale_ratio
                        marinated_items.append((code, new_qty))
                        print(f"    🎯 Found marinated: {MARINATED_CODES[code]} = {new_qty:.2f}kg")
                except (ValueError, TypeError):
                    pass
    
    return marinated_items


# =============================================================================
# CREATE OUTPUT WORKBOOK
# =============================================================================

def create_output_workbook(ws_formula, ws_value, new_batch_size):
    """Create output workbook with formulas and new batch size."""
    
    structure = find_recipe_structure(ws_formula)
    
    # Get column letters for formulas
    control_col_letter = openpyxl.utils.get_column_letter(structure["control_col"])
    percent_col_letter = openpyxl.utils.get_column_letter(structure["percent_col"])
    new_batch_col_letter = openpyxl.utils.get_column_letter(structure["new_batch_col"])
    
    # Create new workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = ws_formula.title
    
    # Determine copy range
    max_row = max(ws_formula.max_row, ws_value.max_row)
    max_col = max(ws_formula.max_column, ws_value.max_column)
    
    # STEP 1: Copy ALL cells
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_formula = ws_formula.cell(r, c)
            src_value = ws_value.cell(r, c)
            tgt_cell = ws_new.cell(r, c)
            
            # Copy VALUE from data_only version
            tgt_cell.value = src_value.value
            
            # Copy STYLING from formula version
            if src_formula.has_style:
                try:
                    tgt_cell.font = copy(src_formula.font)
                    tgt_cell.border = copy(src_formula.border)
                    tgt_cell.fill = copy(src_formula.fill)
                    tgt_cell.number_format = src_formula.number_format
                    tgt_cell.alignment = copy(src_formula.alignment)
                except Exception:
                    pass
    
    # Copy column widths and row heights
    for col, dim in ws_formula.column_dimensions.items():
        ws_new.column_dimensions[col].width = dim.width
    for row, dim in ws_formula.row_dimensions.items():
        ws_new.row_dimensions[row].height = dim.height
    
    # Copy merged cells
    for merged_range in ws_formula.merged_cells.ranges:
        try:
            ws_new.merge_cells(str(merged_range))
        except:
            pass
    
    # STEP 2: Copy print settings
    copy_print_settings(ws_formula, ws_new)
    
    # STEP 3: Update batch size in header
    if structure["batch_size_row"] and structure["batch_size_col"]:
        batch_cell = ws_new.cell(structure["batch_size_row"], structure["batch_size_col"])
        batch_cell.value = new_batch_size
        batch_size_ref = f"${openpyxl.utils.get_column_letter(structure['batch_size_col'])}${structure['batch_size_row']}"
    else:
        print(f"  Warning: Could not find batch size cell")
        batch_size_ref = None
    
    # Pre-compute values needed for steps 4 & 5
    finished_yield = None
    if structure["finished_yield_row"]:
        fy_val = ws_value.cell(structure["finished_yield_row"], structure["control_col"]).value
        try:
            finished_yield = float(fy_val) if fy_val is not None else None
        except (ValueError, TypeError):
            pass

    ingredient_control: dict = {}
    if structure["ingredient_start"] and structure["ingredient_end"]:
        for r in range(structure["ingredient_start"], structure["ingredient_end"] + 1):
            cv = ws_value.cell(r, structure["control_col"]).value
            try:
                ingredient_control[r] = float(cv) if cv is not None else 0.0
            except (ValueError, TypeError):
                ingredient_control[r] = 0.0

    total_in = sum(ingredient_control.values())

    # STEP 4: Write CALCULATED VALUES for ingredient rows
    for r, cv in ingredient_control.items():
        if total_in > 0:
            ws_new.cell(r, structure["percent_col"]).value = round(cv / total_in, 6)
        if finished_yield and finished_yield > 0:
            ws_new.cell(r, structure["new_batch_col"]).value = round(cv * (new_batch_size / finished_yield), 3)

    # STEP 5: Write CALCULATED VALUES for summary rows
    if structure["total_in_row"]:
        ws_new.cell(structure["total_in_row"], structure["control_col"]).value = round(total_in, 3)
        ws_new.cell(structure["total_in_row"], structure["percent_col"]).value = 1.0 if total_in > 0 else 0
        if finished_yield and finished_yield > 0:
            ws_new.cell(structure["total_in_row"], structure["new_batch_col"]).value = round(
                total_in * (new_batch_size / finished_yield), 3
            )

    if structure["finished_yield_row"]:
        fy_control = ws_value.cell(structure["finished_yield_row"], structure["control_col"]).value
        try:
            fy_c = float(fy_control) if fy_control is not None else 0.0
        except (ValueError, TypeError):
            fy_c = 0.0
        if total_in > 0:
            ws_new.cell(structure["finished_yield_row"], structure["percent_col"]).value = round(fy_c / total_in, 6)
        ws_new.cell(structure["finished_yield_row"], structure["new_batch_col"]).value = new_batch_size

    if structure["yield_loss_row"]:
        yl_control = ws_value.cell(structure["yield_loss_row"], structure["control_col"]).value
        try:
            yl_c = float(yl_control) if yl_control is not None else 0.0
        except (ValueError, TypeError):
            yl_c = 0.0
        ws_new.cell(structure["yield_loss_row"], structure["control_col"]).value = round(yl_c, 3)
        if total_in > 0:
            ws_new.cell(structure["yield_loss_row"], structure["percent_col"]).value = round(yl_c / total_in, 6)
        ws_new.cell(structure["yield_loss_row"], structure["new_batch_col"]).value = ""
    
    # STEP 6: HIDE THE CONTROL COLUMN
    ws_new.column_dimensions[control_col_letter].hidden = True
    
    return wb_new, structure


# =============================================================================
# CREATE TEMP RECIPE
# =============================================================================

def create_temp_recipe(master_key: str, sheet: str, batch_size: float, batch_no: int, item_code: str, recipe_map: Dict) -> List[Tuple[str, str, float]]:
    """
    Create temp recipe file and return list of marinated codes found.
    
    Args:
        master_key: The master recipe key (e.g., "KETTLEANDBLENDTECH")
        sheet: The sheet name in the master recipe
        batch_size: The batch size in kg
        batch_no: The batch number
        item_code: The WIP code of this recipe (parent WIP code)
        recipe_map: The recipe mapping dictionary
    
    Returns: List of (parent_wip, marinated_code, quantity) tuples
    """
    print(f"\n{'='*70}")
    print(f"Processing: {sheet}")
    print(f"WIP: {item_code}, Batch #{batch_no}, Target Size: {batch_size} kg")
    print(f"{'='*70}")
    
    # Load BOTH versions of master recipe
    wb_formula = load_master(master_key, data_only=False)
    wb_value = load_master(master_key, data_only=True)
    
    ws_formula = wb_formula[sheet]
    ws_value = wb_value[sheet]
    
    # Create output workbook
    wb_out, structure = create_output_workbook(ws_formula, ws_value, batch_size)
    
    # Save to temp_print folder
    safe_name = re.sub(r"[^\w\s-]", "", sheet).replace(" ", "_")
    filename = f"{safe_name}_BATCH{batch_no}_{int(batch_size)}kg.xlsx"
    path = os.path.join(TEMP_PRINT_FOLDER, filename)
    
    wb_out.save(path)
    print(f"  ✓ Saved: {filename}")
    
    # Get structure and extract marinated codes BEFORE closing workbooks
    # Use ws_value (original master with cached values) to read ingredient codes and control values
    saved_structure = find_recipe_structure(ws_formula)
    marinated_items = extract_marinated_codes_from_recipe(ws_value, saved_structure, batch_size)
    
    # Add parent WIP code to each marinated item
    marinated_with_parent = [(item_code, code, qty) for code, qty in marinated_items]
    
    # Clean up
    wb_formula.close()
    wb_value.close()
    wb_out.close()
    
    return marinated_with_parent


# =============================================================================
# CREATE MARINATION RECIPE
# =============================================================================

def create_marination_recipe(parent_wip: str, marinated_code: str, quantity: float, batch_no: int, recipe_map: Dict):
    """Create marination recipe file for the given marinated code and quantity.
    
    Args:
        parent_wip: The WIP code of the parent recipe (e.g., WBCCCP10000 for Butter Chicken)
        marinated_code: The marinated meat code (e.g., WMDTCCD1000)
        quantity: The quantity needed in kg
        batch_no: The batch number
        recipe_map: The recipe mapping dictionary
    """
    
    if marinated_code not in recipe_map:
        print(f"    ⚠ Marination recipe not found for: {marinated_code}")
        return
    
    recipe_info = recipe_map[marinated_code]
    
    if recipe_info["master"] != "MARINATION":
        print(f"    ⚠ {marinated_code} is not in MARINATION master")
        return
    
    sheet = recipe_info["sheet"]
    
    print(f"\n  → Creating marination recipe: {MARINATED_CODES.get(marinated_code, marinated_code)}")
    print(f"    Parent: {parent_wip}, Sheet: {sheet}, Quantity: {quantity:.2f}kg")
    
    # Load marination master
    wb_formula = load_master("MARINATION", data_only=False)
    wb_value = load_master("MARINATION", data_only=True)
    
    # Check if sheet exists
    if sheet not in wb_formula.sheetnames:
        # Try partial match (Excel 31-char truncation)
        found_sheet = None
        for s in wb_formula.sheetnames:
            if s.lower().startswith(sheet[:25].lower()):
                found_sheet = s
                break
        
        if found_sheet:
            sheet = found_sheet
            print(f"    Using partial match: {sheet}")
        else:
            print(f"    ⚠ Sheet not found: {sheet}")
            wb_formula.close()
            wb_value.close()
            return
    
    ws_formula = wb_formula[sheet]
    ws_value = wb_value[sheet]
    
    # Create output workbook
    wb_out, structure = create_output_workbook(ws_formula, ws_value, quantity)
    
    # Save to temp_print folder - include parent WIP code in filename
    # Format: MARINATION_{parent_wip}_{marinated_code}_BATCH{n}_{qty}kg.xlsx
    filename = f"MARINATION_{parent_wip}_{marinated_code}_BATCH{batch_no}_{int(quantity)}kg.xlsx"
    path = os.path.join(TEMP_PRINT_FOLDER, filename)
    
    wb_out.save(path)
    print(f"    ✓ Saved: {filename}")
    
    # Clean up
    wb_formula.close()
    wb_value.close()
    wb_out.close()


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("BATCH RECIPE PRINTER v2.0")
    print("With Marination Auto-Detection")
    print("=" * 70)
    print(f"Batching file: {BATCHING_FILE}")
    print(f"Output folder: {TEMP_PRINT_FOLDER}")
    print("=" * 70)
    
    # Read batching file
    print("\nReading batching file...")
    entries = read_batching_file()
    print(f"✅ Found {len(entries)} item(s) to process")
    
    # Build recipe map (includes MARINATION)
    print("\nBuilding recipe map...")
    recipe_map = build_master_map()
    print(f"✓ Mapped {len(recipe_map)} recipe(s)")
    
    # Process each entry
    print("\n" + "=" * 70)
    print("CREATING RECIPE FILES")
    print("=" * 70)

    created_count = 0
    marination_count = 0
    skipped_count = 0
    
    # Track all marinated items to create: (parent_wip, marinated_code, qty, batch_no)
    all_marinated_items: List[Tuple[str, str, float, int]] = []
    
    for entry in entries:
        try:
            batch_sizes = parse_batch_size(entry["batch_field"])
            recipe = recipe_map.get(entry["item_code"])
            
            if not recipe:
                print(f"\n⚠ Skipped: {entry['item_code']} (not found in master recipes)")
                skipped_count += 1
                continue
            
            print(f"\n{'='*70}")
            print(f"Item Code: {entry['item_code']}")
            print(f"Total Batches: {len(batch_sizes)}")
            print(f"Total Quantity: {sum(batch_sizes):.0f} kg")
            print(f"{'='*70}")

            # Create a recipe file for each batch
            for i, batch_size in enumerate(batch_sizes, 1):
                try:
                    marinated_items = create_temp_recipe(
                        recipe["master"], 
                        recipe["sheet"], 
                        batch_size, 
                        i,
                        entry["item_code"],  # Pass the parent WIP code
                        recipe_map
                    )
                    created_count += 1
                    
                    # Collect marinated items for later processing
                    # marinated_items is now [(parent_wip, marinated_code, qty), ...]
                    for parent_wip, code, qty in marinated_items:
                        all_marinated_items.append((parent_wip, code, qty, i))
                        
                except Exception as e:
                    print(f"   ❌ Error creating batch {i}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
        except Exception as e:
            print(f"\n❌ Error processing {entry.get('item_code', 'unknown')}: {str(e)}")
            import traceback
            traceback.print_exc()
            skipped_count += 1

    # Create marination recipe files
    if all_marinated_items:
        print("\n" + "=" * 70)
        print("CREATING MARINATION RECIPES")
        print("=" * 70)
        print(f"Found {len(all_marinated_items)} marinated item(s) to expand")
        
        for parent_wip, code, qty, batch_no in all_marinated_items:
            try:
                create_marination_recipe(parent_wip, code, qty, batch_no, recipe_map)
                marination_count += 1
            except Exception as e:
                print(f"   ❌ Error creating marination recipe: {str(e)}")
                import traceback
                traceback.print_exc()

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"✅ Recipe files created: {created_count}")
    print(f"✅ Marination files created: {marination_count}")
    if skipped_count > 0:
        print(f"⚠ Skipped: {skipped_count} item(s)")
    print(f"📁 Output location: {TEMP_PRINT_FOLDER}")
    print("=" * 70)


if __name__ == "__main__":
    main()