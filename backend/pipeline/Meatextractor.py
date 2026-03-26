import os
import logging
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict
import openpyxl
import sys


def resource_path(relative_path):
    """Get path to resource, works for dev and PyInstaller"""
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

TEMP_PRINT_FOLDER = os.path.join(SCRIPT_DIR, "temp_print")
os.makedirs(TEMP_PRINT_FOLDER, exist_ok=True)

MEAT_TEMPLATE = resource_path("Meattemplate.xlsx")
BATCHING_FILE = resource_path("batching_recipes.xlsx")

HEADER_ROW = 6

# RAW MEAT CODES ONLY
MEAT_CODES = {
    "WCKBRDC2020": "Chicken Breast Diced 20x20mm",
    "WBM85IH1000": "Beef Mince 85CL",
    "WBFKNST0010": "Beef Knuckle Strips 10x10mm",
    "RFIMT000054": "Beef Mince 85CL 6mm Fresh",
    "WBM75IH1000": "Beef Mince 75CL",
    "WBPM10000": "Beef Primal Minced",
    "WCKTHIH1000": "Chicken Thigh",
    "WLMIH10000": "Lamb Mince",
    "WCKBRDC1010": "Chicken Diced 10x10mm",
    "WCKWING1000": "Chicken Wings",
    "WBFKNDC1010": "Beef Knuckle In-house Diced 10x10mm",
    "WBFKNDC2020": "Beef Knuckle In-house Diced 20x20mm",
    "WPOMNIH1000": "Pork Mince 85CL In House",
    "WCLDCIH1000": "Chicken Leg Diced 20mm In House",
    "WLLMIHO1000": "Lamb Leg Mince In House",
    "WLAMLEG1000": "Lamb Leg 10mm Dice",
    "WPSSHD10000": "Pork Shoulder 10mm Dice",
    "WBFKNST10000": "Beef Knuckle Strips",
}

# MARINATED MEAT CODES - look for corresponding marination recipe file in temp_print
MARINATED_CODES = {
    "WBMMCP10000": "Beef Marinated",
    "WSNCCA10000": "Chicken Marinated (Singapore)",
    "WMDTCCD1000": "Marinated Diced Tandoori Chicken",
    "WRMCTD10000": "Roasted Moroccan Chicken Thigh 20mm Dices",
    "WLMMCP10000": "Lamb Marinade",
    "WRCTM10000": "Raw Chicken Tikka Marinated",
}

TEMPLATE_COLS = {
    "LINE": 1,
    "PRODUCT": 2,
    "ITEM_CODE": 3,
    "BATCH_SIZE": 4,
    "UOM": 5,
    "MEAT_ITEM": 6,
    "ACTUAL_QTY": 7,
    "INGREDIENT_CODE": 8,
}

TEMPLATE_START_ROW = 5
RECIPE_HEADER_ROW = 6


@dataclass
class MeatQuantity:
    quantity: float
    batch_count: int


@dataclass
class MeatTypeData:
    code: str
    quantities: List[MeatQuantity]

    def to_string(self) -> str:
        parts = []
        for mq in self.quantities:
            parts.append(
                f"{mq.quantity:.2f}*{mq.batch_count}"
                if mq.batch_count > 1
                else f"{mq.quantity:.2f}"
            )
        return f"({' / '.join(parts)})" if len(parts) > 1 else parts[0]


@dataclass
class WIPMeatData:
    meat_types: List[MeatTypeData]
    batch_sizes: List[Tuple[float, int]]

    def to_string(self) -> Tuple[str, str]:
        qty = " / ".join(mt.to_string() for mt in self.meat_types)
        codes = " / ".join(mt.code for mt in self.meat_types)
        return qty, codes

    def get_batch_size_display(self) -> str:
        return "/".join(
            f"{b:.0f}*{c}" if c > 1 else f"{b:.0f}"
            for b, c in self.batch_sizes
        )


def setup_logging() -> logging.Logger:
    log_folder = os.path.join(SCRIPT_DIR, "logs")
    os.makedirs(log_folder, exist_ok=True)

    logger = logging.getLogger("MeatExtractor")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fh = logging.FileHandler(os.path.join(log_folder, f"meat_{ts}.log"), encoding="utf-8")
    fh.setLevel(logging.DEBUG)

    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)  # Changed to DEBUG to see more output

    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    fh.setFormatter(fmt)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)

    return logger


def normalize_text(val) -> str:
    return str(val).strip().lower() if val else ""


# =============================================================================
# FIND RECIPE STRUCTURE
# =============================================================================

def find_recipe_structure(ws, logger: logging.Logger) -> dict:
    """Find column positions and key cells in recipe."""
    
    ingredient_code_col = None
    ingredient_desc_col = None
    control_col = None
    new_batch_col = None
    
    for col in range(1, 20):
        header = ws.cell(HEADER_ROW, col).value
        if not header:
            continue
        text = normalize_text(header)
        
        if "ingredient" in text and "code" in text:
            ingredient_code_col = col
        elif "ingredient" in text and ("description" in text or "desc" in text):
            ingredient_desc_col = col
        elif text == "control":
            control_col = col
        elif "new" in text and "batch" in text:
            new_batch_col = col
    
    # Find ingredient row range
    ingredient_start = HEADER_ROW + 1
    ingredient_end = None
    
    search_col = ingredient_desc_col or control_col or 2
    for r in range(ingredient_start, min(ingredient_start + 50, ws.max_row + 1)):
        cell_val = ws.cell(r, search_col).value
        if cell_val:
            text = normalize_text(str(cell_val))
            if "total" in text and "in" in text:
                ingredient_end = r - 1
                break
    
    if not ingredient_end or ingredient_end < ingredient_start:
        ingredient_end = ingredient_start + 15
    
    # Find finished yield row for calculations
    finished_yield_row = None
    for r in range(ingredient_end, min(ingredient_end + 10, ws.max_row + 1)):
        cell_val = ws.cell(r, search_col).value
        if cell_val:
            text = normalize_text(str(cell_val))
            if "finished" in text and "yield" in text:
                finished_yield_row = r
                break
    
    return {
        "ingredient_code_col": ingredient_code_col,
        "ingredient_desc_col": ingredient_desc_col,
        "control_col": control_col,
        "new_batch_col": new_batch_col,
        "ingredient_start": ingredient_start,
        "ingredient_end": ingredient_end,
        "finished_yield_row": finished_yield_row,
    }


# =============================================================================
# FIND WIP CODE IN RECIPE
# =============================================================================

def find_wip_code_in_recipe(ws, logger: logging.Logger) -> Optional[str]:
    """
    Find WIP code in recipe file.
    Search common locations in header area.
    """
    # Common locations for WIP code
    locations_to_check = [
        (4, 6),  # Row 4, Col F - original location
        (4, 5),  # Row 4, Col E
        (3, 6),  # Row 3, Col F
        (5, 6),  # Row 5, Col F
        (4, 4),  # Row 4, Col D
    ]
    
    for row, col in locations_to_check:
        val = ws.cell(row, col).value
        if val:
            code = str(val).strip().upper()
            # Check if it looks like a WIP code (starts with W or R, has numbers)
            if (code.startswith('W') or code.startswith('R')) and any(c.isdigit() for c in code):
                logger.debug(f"    Found WIP code at ({row}, {col}): {code}")
                return code
    
    # Search header area more broadly
    for r in range(1, 6):
        for c in range(1, 10):
            val = ws.cell(r, c).value
            if val:
                code = str(val).strip().upper()
                if (code.startswith('W') or code.startswith('R')) and len(code) >= 8 and any(c.isdigit() for c in code):
                    # Verify it looks like a WIP code pattern
                    if re.match(r'^[WR][A-Z0-9]{7,}$', code):
                        logger.debug(f"    Found WIP code at ({r}, {c}): {code}")
                        return code
    
    return None


# =============================================================================
# FIND MARINATION FILE IN TEMP_PRINT
# =============================================================================

def find_marination_file(parent_wip: str, marinated_code: str, batch_no: int, logger: logging.Logger) -> Optional[str]:
    """
    Find the marination recipe file in temp_print folder.
    Files are named: MARINATION_{parent_wip}_{marinated_code}_BATCH{n}_{qty}kg.xlsx
    
    Args:
        parent_wip: The WIP code of the parent recipe (e.g., WBCCCP10000)
        marinated_code: The marinated meat code (e.g., WMDTCCD1000)
        batch_no: The batch number
        logger: Logger instance
    """
    
    # Get all marination files
    marination_files = [
        f for f in os.listdir(TEMP_PRINT_FOLDER)
        if f.startswith('MARINATION_')
        and f.endswith('.xlsx')
        and not f.startswith('~$')
    ]
    
    if not marination_files:
        logger.debug(f"    No marination files found in temp_print")
        return None
    
    logger.debug(f"    Looking for: MARINATION_{parent_wip}_{marinated_code}_BATCH{batch_no}_*.xlsx")
    logger.debug(f"    Available marination files: {marination_files}")
    
    # Try to find exact match by parent WIP, marinated code, and batch number
    for filename in marination_files:
        if (f"MARINATION_{parent_wip}_{marinated_code}_BATCH{batch_no}_" in filename):
            filepath = os.path.join(TEMP_PRINT_FOLDER, filename)
            logger.info(f"    ✓ Found marination file: {filename}")
            return filepath
    
    # If no exact batch match, try matching just parent WIP and marinated code
    for filename in marination_files:
        if (f"MARINATION_{parent_wip}_{marinated_code}_" in filename):
            filepath = os.path.join(TEMP_PRINT_FOLDER, filename)
            logger.info(f"    ✓ Found marination file (partial match): {filename}")
            return filepath
    
    # Fallback: try matching just marinated code
    for filename in marination_files:
        if f"_{marinated_code}_" in filename:
            filepath = os.path.join(TEMP_PRINT_FOLDER, filename)
            logger.info(f"    ✓ Found marination file (code match): {filename}")
            return filepath
    
    logger.warning(f"    No matching marination file for {parent_wip} -> {marinated_code}")
    return None


# =============================================================================
# EXTRACT RAW MEAT FROM MARINATION FILE
# =============================================================================

def extract_raw_meat_from_marination(filepath: str, logger: logging.Logger) -> Dict[str, float]:
    """
    Extract raw meat codes and quantities from a marination recipe file.
    Uses the Control column and calculates scaled quantities based on batch size.
    """
    filename = os.path.basename(filepath)
    logger.info(f"    Reading marination file: {filename}")
    print(f"    [DEBUG] Reading marination file: {filename}")
    
    try:
        # Load with data_only=True to get values (not formulas) from Control column
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # For marination files, we know the structure:
        # Header row 6, ingredients start row 7
        # Col B (2) = Ingredient Code
        # Col D (4) = Control
        # Col F (6) = New Batch
        
        code_col = 2  # Column B
        control_col = 4  # Column D
        
        # Find finished yield row (look for "Finished yield" in column C)
        finished_yield_row = None
        finished_yield = None
        for r in range(15, 30):
            cell_val = ws.cell(r, 3).value  # Column C
            if cell_val and "finished" in str(cell_val).lower():
                finished_yield_row = r
                fy_val = ws.cell(r, control_col).value
                if fy_val:
                    try:
                        finished_yield = float(fy_val)
                    except:
                        pass
                break
        
        print(f"    [DEBUG] Finished yield row: {finished_yield_row}, value: {finished_yield}")
        
        # Get batch size from filename
        batch_size = get_batch_size_from_filename(filename)
        print(f"    [DEBUG] Batch size from filename: {batch_size}")
        
        # Calculate scale ratio
        if batch_size and finished_yield and finished_yield > 0:
            scale_ratio = batch_size / finished_yield
            print(f"    [DEBUG] Scale: {batch_size}kg / {finished_yield}kg = {scale_ratio:.4f}")
        else:
            scale_ratio = 1.0
            print(f"    [DEBUG] Could not calculate scale ratio, using 1.0")
        
        raw_meat_dict = {}
        
        # Scan rows 7 to finished_yield_row - 2 (to skip Total in row)
        end_row = (finished_yield_row - 2) if finished_yield_row else 25
        print(f"    [DEBUG] Scanning rows 7 to {end_row}")
        
        for r in range(7, end_row + 1):
            code_val = ws.cell(r, code_col).value
            ctrl_val = ws.cell(r, control_col).value
            
            if not code_val:
                continue
            
            code = str(code_val).strip().upper()
            
            # Log every row
            is_meat = code in MEAT_CODES
            print(f"    [DEBUG] Row {r}: code='{code}', control={ctrl_val}, is_meat={is_meat}")
            
            # Check if this is RAW MEAT
            if is_meat:
                if ctrl_val is not None:
                    try:
                        ctrl_qty = float(ctrl_val)
                        qty = ctrl_qty * scale_ratio
                        if qty > 0:
                            raw_meat_dict[code] = raw_meat_dict.get(code, 0) + qty
                            print(f"    [DEBUG] ✓ FOUND RAW MEAT: {MEAT_CODES[code]}: {ctrl_qty:.2f} * {scale_ratio:.4f} = {qty:.2f}kg")
                            logger.info(f"      ✓ {MEAT_CODES[code]}: {ctrl_qty:.2f} * {scale_ratio:.4f} = {qty:.2f}kg")
                    except (ValueError, TypeError) as e:
                        print(f"    [DEBUG] Could not convert {ctrl_val} to float: {e}")
        
        wb.close()
        
        if raw_meat_dict:
            print(f"    [DEBUG] ✅ Extracted {len(raw_meat_dict)} raw meat items: {raw_meat_dict}")
            logger.info(f"    ✅ Extracted {len(raw_meat_dict)} raw meat items")
        else:
            print(f"    [DEBUG] ⚠️ No raw meat found!")
            logger.warning(f"    ⚠️ No raw meat found in marination file")
        
        return raw_meat_dict
    
    except Exception as e:
        print(f"    [DEBUG] ❌ Error: {e}")
        logger.error(f"    ❌ Error reading marination file: {e}")
        import traceback
        traceback.print_exc()
        logger.error(traceback.format_exc())
        return {}


# =============================================================================
# READ BATCH INFO
# =============================================================================

def read_batch_info(logger: logging.Logger) -> Dict[str, str]:
    """Read batch information from batching_recipes.xlsx"""
    logger.info("Reading batch information...")
    
    if not os.path.exists(BATCHING_FILE):
        logger.warning(f"Batching file not found: {BATCHING_FILE}")
        return {}
    
    try:
        wb = openpyxl.load_workbook(BATCHING_FILE, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                key = normalize_text(header).replace(" ", "")
                headers[key] = col
        
        wip_col = None
        batch_col = None
        
        for key in ["itemcode", "wipcode", "item"]:
            if key in headers:
                wip_col = headers[key]
                break
        
        for key in ["batchsize", "batch", "size"]:
            if key in headers:
                batch_col = headers[key]
                break
        
        if not wip_col or not batch_col:
            wb.close()
            return {}
        
        batch_info = {}
        for row in range(2, ws.max_row + 1):
            wip = ws.cell(row, wip_col).value
            batch = ws.cell(row, batch_col).value
            if wip and batch:
                batch_info[str(wip).strip()] = str(batch)
        
        wb.close()
        logger.info(f"✓ Loaded {len(batch_info)} batch records")
        return batch_info
    
    except Exception as e:
        logger.error(f"Error reading batching file: {e}")
        return {}


# =============================================================================
# GET BATCH INFO FROM FILENAME
# =============================================================================

def get_batch_size_from_filename(filename: str) -> Optional[float]:
    """Extract batch size from filename like 'RECIPE_BATCH1_2000kg.xlsx'"""
    match = re.search(r'_(\d+(?:\.\d+)?)kg', filename, re.IGNORECASE)
    if match:
        return float(match.group(1))
    return None


def get_batch_number_from_filename(filename: str) -> int:
    """Extract batch number from filename like 'RECIPE_BATCH1_2000kg.xlsx'"""
    match = re.search(r'_BATCH(\d+)_', filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 1


# =============================================================================
# EXTRACT MEAT FROM RECIPE
# =============================================================================

def extract_meat_from_recipe(filepath: str, logger: logging.Logger) -> tuple:
    """Extract WIP code, batch size, and meat data from a recipe file"""
    filename = os.path.basename(filepath)
    
    try:
        wb_data = openpyxl.load_workbook(filepath, data_only=True)
        ws_data = wb_data.active
        
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
        ws_formulas = wb_formulas.active
        
        # Get WIP code - try multiple methods
        wip_code = find_wip_code_in_recipe(ws_data, logger)
        if not wip_code:
            wip_code = find_wip_code_in_recipe(ws_formulas, logger)
        
        if not wip_code:
            logger.warning(f"  ⚠️ No WIP code found")
            # Debug: show what's in the header area
            logger.debug(f"  Header area contents:")
            for r in range(1, 7):
                row_vals = [ws_data.cell(r, c).value for c in range(1, 10)]
                logger.debug(f"    Row {r}: {row_vals}")
            wb_data.close()
            wb_formulas.close()
            return None, None, None
        
        logger.info(f"  WIP Code: {wip_code}")
        
        batch_size = get_batch_size_from_filename(filename)
        if not batch_size:
            logger.warning(f"  ⚠️ No batch size in filename")
            wb_data.close()
            wb_formulas.close()
            return wip_code, None, None
        
        batch_no = get_batch_number_from_filename(filename)
        
        # Find recipe structure
        structure = find_recipe_structure(ws_formulas, logger)
        
        logger.debug(f"  Structure: code_col={structure['ingredient_code_col']}, control_col={structure['control_col']}, new_batch_col={structure['new_batch_col']}")
        logger.debug(f"  Ingredients: rows {structure['ingredient_start']}-{structure['ingredient_end']}")
        
        if not structure["ingredient_code_col"]:
            logger.warning(f"  ⚠️ Could not find Ingredient Code column")
            wb_data.close()
            wb_formulas.close()
            return wip_code, batch_size, None
        
        code_col = structure["ingredient_code_col"]
        control_col = structure["control_col"]
        new_batch_col = structure["new_batch_col"]
        
        # Get finished yield for scaling calculations
        finished_yield = None
        if structure["finished_yield_row"] and control_col:
            fy_val = ws_data.cell(structure["finished_yield_row"], control_col).value
            if fy_val:
                try:
                    finished_yield = float(fy_val)
                except (ValueError, TypeError):
                    pass
        
        # Calculate scale ratio if we have finished yield
        scale_ratio = batch_size / finished_yield if finished_yield and finished_yield > 0 else 1.0
        
        logger.debug(f"  Finished yield: {finished_yield}, Scale ratio: {scale_ratio:.4f}")
        
        meat_dict = {}
        ingredients_found = []
        
        for row in range(structure["ingredient_start"], structure["ingredient_end"] + 1):
            # Get code
            code_value = ws_data.cell(row, code_col).value or ws_formulas.cell(row, code_col).value
            if not code_value:
                continue
            
            code = str(code_value).strip().upper()
            
            # Skip summary rows
            if any(word in code.lower() for word in ["total", "finished", "yield", "loss"]):
                break
            
            # Get quantity - try New Batch column first, then Control column with scaling
            qty = None
            
            # Try New Batch column (might have cached values)
            if new_batch_col:
                nb_val = ws_data.cell(row, new_batch_col).value
                if nb_val is not None:
                    try:
                        qty = float(nb_val)
                    except (ValueError, TypeError):
                        pass
            
            # If no New Batch value, use Control column with scaling
            if qty is None and control_col:
                ctrl_val = ws_data.cell(row, control_col).value
                if ctrl_val is not None:
                    try:
                        qty = float(ctrl_val) * scale_ratio
                    except (ValueError, TypeError):
                        pass
            
            ingredients_found.append((code, qty))
            
            # Check if RAW MEAT - directly extract
            if code in MEAT_CODES:
                if qty and qty > 0:
                    meat_dict[code] = meat_dict.get(code, 0) + qty
                    logger.info(f"  ✓ RAW MEAT: {MEAT_CODES[code]} = {qty:.2f}kg")
            
            # Check if MARINATED MEAT - find marination file and extract
            elif code in MARINATED_CODES:
                if qty and qty > 0:
                    logger.info(f"  🎯 Found marinated: {MARINATED_CODES[code]} ({qty:.2f}kg)")
                    
                    # Find the corresponding marination file using parent WIP code
                    marination_file = find_marination_file(wip_code, code, batch_no, logger)
                    
                    if marination_file:
                        raw_meats = extract_raw_meat_from_marination(marination_file, logger)
                        for raw_code, raw_qty in raw_meats.items():
                            meat_dict[raw_code] = meat_dict.get(raw_code, 0) + raw_qty
                    else:
                        logger.warning(f"    ⚠️ No marination file found - cannot expand {code}")
        
        # Debug: show what ingredients were found (always show for debugging)
        if ingredients_found:
            logger.debug(f"  All ingredients found ({len(ingredients_found)}):")
            for code, qty in ingredients_found[:15]:
                marker = ""
                if code in MEAT_CODES:
                    marker = " ← RAW MEAT"
                elif code in MARINATED_CODES:
                    marker = " ← MARINATED"
                logger.debug(f"    {code} = {qty}{marker}")
            if len(ingredients_found) > 15:
                logger.debug(f"    ... and {len(ingredients_found) - 15} more")
        
        if not meat_dict:
            logger.warning(f"  ⚠️ No meat found in recipe")
        
        wb_data.close()
        wb_formulas.close()
        
        return wip_code, batch_size, meat_dict if meat_dict else None
    
    except Exception as e:
        logger.error(f"  ❌ Error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None, None, None


# =============================================================================
# SCAN ALL RECIPES
# =============================================================================

def scan_all_recipes(logger: logging.Logger) -> Dict[str, WIPMeatData]:
    """Scan all recipe files and group by WIP code"""
    logger.info("=" * 70)
    logger.info("SCANNING RECIPE FILES")
    logger.info("=" * 70)
    
    if not os.path.exists(TEMP_PRINT_FOLDER):
        logger.error(f"Folder not found: {TEMP_PRINT_FOLDER}")
        return {}
    
    # Get recipe files (exclude marination files - they are processed via main recipes)
    recipe_files = [
        f for f in os.listdir(TEMP_PRINT_FOLDER)
        if f.endswith('.xlsx')
        and 'BATCH' in f.upper()
        and 'MEAT_PRODUCTION' not in f.upper()
        and not f.startswith('MARINATION_')  # Skip marination files - processed separately
        and not f.startswith('~$')
    ]
    
    if not recipe_files:
        logger.error("No recipe files found")
        return {}
    
    # Also check for marination files
    marination_files = [f for f in os.listdir(TEMP_PRINT_FOLDER) if f.startswith('MARINATION_')]
    logger.info(f"Found {len(recipe_files)} recipe files")
    logger.info(f"Found {len(marination_files)} marination files")
    logger.info("")
    
    wip_batch_map = defaultdict(lambda: defaultdict(list))
    
    for filename in sorted(recipe_files):
        filepath = os.path.join(TEMP_PRINT_FOLDER, filename)
        logger.info(f"Processing: {filename}")
        
        wip_code, batch_size, meat_dict = extract_meat_from_recipe(filepath, logger)
        
        if wip_code and batch_size and meat_dict:
            wip_batch_map[wip_code][batch_size].append(meat_dict)
            logger.info(f"  ✓ {wip_code}: {batch_size}kg, {len(meat_dict)} meat items")
        elif wip_code and batch_size:
            logger.warning(f"  ⚠️ {wip_code}: {batch_size}kg - NO MEAT FOUND")
        logger.info("")
    
    result = {}
    for wip_code, batch_sizes_dict in wip_batch_map.items():
        sorted_sizes = sorted(batch_sizes_dict.keys())
        
        all_codes = set()
        for dicts in batch_sizes_dict.values():
            for d in dicts:
                all_codes.update(d.keys())
        
        meat_types = []
        for code in sorted(all_codes):
            quantities = []
            for size in sorted_sizes:
                batch_list = batch_sizes_dict[size]
                qty = batch_list[0].get(code, 0)
                if qty > 0:
                    quantities.append(MeatQuantity(qty, len(batch_list)))
            
            if quantities:
                meat_types.append(MeatTypeData(code, quantities))
        
        result[wip_code] = WIPMeatData(
            meat_types=meat_types,
            batch_sizes=[(s, len(batch_sizes_dict[s])) for s in sorted_sizes]
        )
    
    logger.info(f"Found meat data for {len(result)} WIP codes")
    return result


# =============================================================================
# UPDATE MEAT TEMPLATE
# =============================================================================

def update_meat_template(
    meat_by_wip: Dict[str, WIPMeatData],
    batch_info: Dict[str, str],
    logger: logging.Logger
) -> Optional[str]:
    logger.info("=" * 70)
    logger.info("UPDATING TEMPLATE")
    logger.info("=" * 70)
    
    if not os.path.exists(MEAT_TEMPLATE):
        logger.error(f"Template not found: {MEAT_TEMPLATE}")
        return None
    
    try:
        wb = openpyxl.load_workbook(MEAT_TEMPLATE)
        ws = wb.active
        
        updated = 0
        hidden = 0
        
        for row in range(TEMPLATE_START_ROW, ws.max_row + 1):
            wip = ws.cell(row, TEMPLATE_COLS["ITEM_CODE"]).value
            wip = str(wip).strip() if wip else None
            
            if not wip or wip not in meat_by_wip:
                ws.row_dimensions[row].hidden = True
                hidden += 1
                continue
            
            data = meat_by_wip[wip]
            
            ws.cell(row, TEMPLATE_COLS["BATCH_SIZE"]).value = batch_info.get(wip, data.get_batch_size_display())
            
            qty_str, code_str = data.to_string()
            ws.cell(row, TEMPLATE_COLS["ACTUAL_QTY"]).value = qty_str
            ws.cell(row, TEMPLATE_COLS["INGREDIENT_CODE"]).value = code_str
            
            updated += 1
            product = ws.cell(row, TEMPLATE_COLS["PRODUCT"]).value or "Unknown"
            logger.info(f"✓ {product}: {qty_str}")
        
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        output = os.path.join(TEMP_PRINT_FOLDER, f"MEAT_PRODUCTION_LIST_{ts}.xlsx")
        
        wb.save(output)
        wb.close()
        
        logger.info("")
        logger.info(f"Updated: {updated} rows, Hidden: {hidden} rows")
        logger.info(f"Output: {output}")
        
        return output
    
    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


# =============================================================================
# MAIN
# =============================================================================

def main():
    logger = setup_logging()
    
    logger.info("=" * 70)
    logger.info("MEAT ROOM PRODUCTION LIST GENERATOR")
    logger.info("Version 9.1.0 - Marination File Integration")
    logger.info("=" * 70)
    logger.info(f"Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("")
    
    try:
        batch_info = read_batch_info(logger)
        logger.info("")
        
        meat_by_wip = scan_all_recipes(logger)
        
        if not meat_by_wip:
            logger.error("No meat data found!")
            logger.info("")
            logger.info("Troubleshooting tips:")
            logger.info("1. Check that recipe files contain meat codes in MEAT_CODES dictionary")
            logger.info("2. Check that WIP codes can be found in recipe files")
            logger.info("3. Run batch printer first to generate marination files")
            return
        
        logger.info("")
        output = update_meat_template(meat_by_wip, batch_info, logger)
        
        logger.info("")
        logger.info("=" * 70)
        if output:
            logger.info("✅ SUCCESS")
            logger.info(f"📄 {os.path.basename(output)}")
        else:
            logger.error("❌ FAILED")
        logger.info("=" * 70)
    
    except Exception as e:
        logger.error(f"Critical error: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    logger.info(f"End: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()