"""
FastAPI backend for the Recipe Automation system.
Orchestrates the Python pipeline, writes results to Supabase, and
exposes download endpoints for the Next.js frontend.

Required env vars:
  SUPABASE_URL          – project URL (https://xxx.supabase.co)
  SUPABASE_SERVICE_KEY  – service-role key (kept server-side only)
  RECIPE_AUTOMATION_DIR – path to the folder containing the Python scripts
                          (default: backend/pipeline/ relative to this file)
"""

from __future__ import annotations

import asyncio
import io
import os
import re
import subprocess
import sys
import uuid
import zipfile
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

import httpx
import openpyxl
from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from supabase import Client, create_client

load_dotenv()  # reads backend/.env

# ─────────────────────────────── Config ──────────────────────────────────────

SUPABASE_URL: str        = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY: str = os.environ["SUPABASE_SERVICE_KEY"]
WEBHOOK_SECRET: str       = os.environ.get("WEBHOOK_SECRET", "")
# On Railway: set RECIPE_AUTOMATION_DIR env var, e.g. /app/pipeline
# Locally: defaults to the pipeline/ subfolder next to this file
RECIPE_AUTOMATION_DIR = Path(
    os.environ.get(
        "RECIPE_AUTOMATION_DIR",
        str(Path(__file__).parent / "pipeline"),
    )
)
MASTER_BUCKET  = "master-recipes"
STORAGE_BUCKET = "recipe-outputs"
HEADER_ROW = 6  # Row 6 contains column headers in every recipe xlsx

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# ─────────────────────── Filename patterns ───────────────────────────────────

# Butter_Chicken_BATCH1_1800kg.xlsx
_RE_RECIPE = re.compile(r"^(.+)_BATCH(\d+)_(\d+)kg\.xlsx$", re.IGNORECASE)

# MARINATION_WBSCCP1000_WBMMCP10000_BATCH1_500kg.xlsx
_RE_MARINATION = re.compile(
    r"^MARINATION_(\w+)_(\w+)_BATCH(\d+)_(\d+)kg\.xlsx$", re.IGNORECASE
)

# ─────────────────── Startup: download master recipe files ───────────────────

MASTER_FILES = [
    "Recipes/KETTLEANDBLENDTECH.xlsx",
    "Recipes/SOUP.xlsx",
    "Recipes/SAUCE.xlsx",
    "Recipes/PASTRY.xlsx",
    "Recipes/MARINATION.xlsx",
    "Recipes/WOK.xlsx",
    "Recipes/RICE.xlsx",
    "Recipes/CQC.xlsx",
]


async def download_master_files() -> None:
    """
    Download password-protected master recipe files from Supabase Storage
    into RECIPE_AUTOMATION_DIR/Recipes/ so the pipeline scripts can find them.
    Skips files that already exist locally (avoids re-downloading on restart).
    """
    recipes_dir = RECIPE_AUTOMATION_DIR / "Recipes"
    recipes_dir.mkdir(parents=True, exist_ok=True)

    for rel_path in MASTER_FILES:
        dest = RECIPE_AUTOMATION_DIR / rel_path
        if dest.exists():
            continue  # already downloaded
        try:
            data = supabase.storage.from_(MASTER_BUCKET).download(rel_path)
            dest.write_bytes(data)
            print(f"[startup] Downloaded {rel_path}")
        except Exception as exc:
            print(f"[startup] WARNING: could not download {rel_path}: {exc}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Ensure pipeline directory and temp_print exist
    (RECIPE_AUTOMATION_DIR / "temp_print").mkdir(parents=True, exist_ok=True)
    await download_master_files()
    yield


# ─────────────────────────────── App ─────────────────────────────────────────

app = FastAPI(title="Recipe Automation API", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────── Subprocess helper ──────────────────────────────────


async def _run_script(script_path: Path, cwd: Path) -> None:
    """Run a Python script in a thread executor (works on Windows + Linux)."""
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"  # prevent Unicode errors on Windows cp1252
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(cwd),
            capture_output=True,
            text=True,
            encoding="utf-8",
            env=env,
        ),
    )
    if result.returncode != 0:
        raise RuntimeError(f"{script_path.name} failed:\n{result.stderr}")


# ──────────────────────── Pipeline models ────────────────────────────────────


class RunRequest(BaseModel):
    production_file: str  # path / URL to the production xlsx
    notes: Optional[str] = None


class RunStatus(BaseModel):
    run_id: str
    status: str  # queued | running | done | error
    message: Optional[str] = None


# ──────────────────────── Helpers ────────────────────────────────────────────


def _norm(val) -> str:
    return str(val).strip().lower() if val else ""


def _parse_recipe_file(path: Path) -> list[dict]:
    """
    Extract ingredient rows from a recipe batch xlsx.

    Mirrors RecipeBatchPrinter.py's find_recipe_structure() — scans row 6 to
    locate each column dynamically, then reads rows 7+ until the first summary
    row ('Total In'). Columns may appear in different orders across master files.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # ── Step 1: find column indices from header row (row 6, 1-based) ─────────
    mix_col = ing_code_col = desc_col = pct_col = new_batch_col = None

    for col in range(1, ws.max_column + 1):
        h = _norm(ws.cell(HEADER_ROW, col).value)
        if not h:
            continue
        if h == "mix":
            mix_col = col
        elif "ingredient" in h and "code" in h:
            ing_code_col = col
        elif "ingredient" in h and ("description" in h or "desc" in h):
            desc_col = col
        elif h == "%":
            pct_col = col
        elif "new" in h and "batch" in h:
            new_batch_col = col

    # ── Step 2: read data rows, stop at summary rows ──────────────────────────
    rows: list[dict] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        # Stop at "Total In" / "Finished Yield" summary rows
        desc_val = ws.cell(r, desc_col).value if desc_col else None
        if desc_val:
            t = _norm(str(desc_val))
            if "total" in t and "in" in t:
                break

        mix      = ws.cell(r, mix_col).value      if mix_col      else None
        ing_code = ws.cell(r, ing_code_col).value if ing_code_col else None
        desc     = ws.cell(r, desc_col).value     if desc_col     else None
        pct      = ws.cell(r, pct_col).value      if pct_col      else None
        new_qty  = ws.cell(r, new_batch_col).value if new_batch_col else None

        if not any([mix, ing_code, desc, pct, new_qty]):
            continue

        rows.append(
            {
                "mix":              str(mix).strip()      if mix      is not None else None,
                "ingredient_code":  str(ing_code).strip() if ing_code is not None else None,
                "description":      str(desc).strip()     if desc     is not None else None,
                "percentage":       float(pct)     if pct     is not None else None,
                "new_batch_qty":    float(new_qty)  if new_qty  is not None else None,
            }
        )

    wb.close()
    return rows


def _classify_file(filename: str) -> Optional[dict]:
    """
    Return metadata dict for a recipe xlsx filename, or None if unrecognised.
    """
    m = _RE_MARINATION.match(filename)
    if m:
        parent_wip, _marinated_code, batch_num, size_kg = m.groups()
        return {
            "file_type": "marination",
            "item_code": parent_wip,
            "batch_number": int(batch_num),
            "batch_size_kg": int(size_kg),
        }

    m = _RE_RECIPE.match(filename)
    if m:
        safe_name, batch_num, size_kg = m.groups()
        return {
            "file_type": "recipe",
            "item_code": safe_name,
            "batch_number": int(batch_num),
            "batch_size_kg": int(size_kg),
        }

    return None


# ──────────────────────── Pipeline ───────────────────────────────────────────


async def run_pipeline(run_id: str, production_file: str) -> None:
    """
    Full pipeline for one run.

    Step 1 – run Python automation scripts
    Step 2 – upload xlsx outputs to Supabase Storage + parse ingredients
    """
    temp_dir = RECIPE_AUTOMATION_DIR / "temp_print"

    try:
        # ── Update run status → running ──────────────────────────────────────
        supabase.table("runs").update({"status": "running"}).eq("id", run_id).execute()

        # ── Step 1: Run automation scripts ───────────────────────────────────
        scripts = ["Read.py", "RecipeBatchPrinter.py", "Meatextractor.py", "templeteeditor.py"]
        for script in scripts:
            script_path = RECIPE_AUTOMATION_DIR / script
            if not script_path.exists():
                continue
            await _run_script(script_path, RECIPE_AUTOMATION_DIR)

        # ── Step 2: Upload files + parse ingredients ──────────────────────────
        if temp_dir.exists():
            recipe_file_records: list[dict] = []
            ingredient_records: list[dict] = []

            for xlsx_path in sorted(temp_dir.glob("*.xlsx")):
                filename = xlsx_path.name
                meta = _classify_file(filename)
                if meta is None:
                    continue  # unrecognised filename — skip

                storage_path = f"{run_id}/{filename}"

                # Upload to Supabase Storage
                with open(xlsx_path, "rb") as fh:
                    supabase.storage.from_(STORAGE_BUCKET).upload(
                        storage_path,
                        fh.read(),
                        {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "upsert": "true"},
                    )

                # record for recipe_files table
                recipe_file_records.append(
                    {
                        "run_id": run_id,
                        "file_name": filename,
                        "storage_path": storage_path,
                        **meta,
                    }
                )

                # Parse ingredient rows
                rows = _parse_recipe_file(xlsx_path)
                for row in rows:
                    ingredient_records.append(
                        {
                            "run_id": run_id,
                            "file_name": filename,
                            "item_code": meta["item_code"],
                            "file_type": meta["file_type"],
                            "batch_number": meta["batch_number"],
                            "batch_size_kg": meta["batch_size_kg"],
                            **row,
                        }
                    )

            # Batch-insert recipe_files
            if recipe_file_records:
                supabase.table("recipe_files").insert(recipe_file_records).execute()

            # Batch-insert recipe_ingredients
            if ingredient_records:
                supabase.table("recipe_ingredients").insert(ingredient_records).execute()

            # ── Clean up temp_print so the next run starts fresh ──────────────
            for f in temp_dir.glob("*.xlsx"):
                f.unlink(missing_ok=True)

        # ── Mark run as done ──────────────────────────────────────────────────
        supabase.table("runs").update({"status": "done"}).eq("id", run_id).execute()

    except Exception as exc:
        supabase.table("runs").update(
            {"status": "error", "error_message": str(exc)}
        ).eq("id", run_id).execute()
        raise


# ──────────────────────── Routes ─────────────────────────────────────────────


@app.post("/api/runs", response_model=RunStatus, status_code=202)
async def create_run(body: RunRequest, background_tasks: BackgroundTasks) -> RunStatus:
    """Kick off a new pipeline run in the background."""
    run_id = str(uuid.uuid4())
    supabase.table("runs").insert(
        {"id": run_id, "status": "queued", "notes": body.notes}
    ).execute()
    background_tasks.add_task(run_pipeline, run_id, body.production_file)
    return RunStatus(run_id=run_id, status="queued")


@app.get("/api/runs/{run_id}", response_model=RunStatus)
async def get_run(run_id: str) -> RunStatus:
    """Poll the status of a run."""
    result = supabase.table("runs").select("*").eq("id", run_id).maybe_single().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="Run not found")
    row = result.data
    return RunStatus(run_id=run_id, status=row["status"], message=row.get("error_message"))


@app.get("/api/runs/{run_id}/recipes/zip")
async def download_recipes_zip(run_id: str) -> StreamingResponse:
    """
    Build and stream a ZIP of all recipe xlsx files for the given run.

    Archive layout:
      Recipes/      ← file_type = 'recipe'
      Marination/   ← file_type = 'marination'
    """
    # Fetch recipe_files rows for this run
    result = (
        supabase.table("recipe_files")
        .select("file_name, storage_path, file_type")
        .eq("run_id", run_id)
        .execute()
    )
    files = result.data or []
    if not files:
        raise HTTPException(status_code=404, detail="No recipe files found for this run")

    # Download files and build in-memory ZIP
    zip_buffer = io.BytesIO()
    async with httpx.AsyncClient(timeout=60) as client:
        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for file_row in files:
                # Generate a short-lived signed URL
                signed = supabase.storage.from_(STORAGE_BUCKET).create_signed_url(
                    file_row["storage_path"], expires_in=120
                )
                signed_url: str = signed["signedURL"]

                # Download file content
                resp = await client.get(signed_url)
                if resp.status_code != 200:
                    continue  # skip unavailable files rather than aborting

                folder = "Marination" if file_row["file_type"] == "marination" else "Recipes"
                arc_name = f"{folder}/{file_row['file_name']}"
                zf.writestr(arc_name, resp.content)

    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="recipes_{run_id[:8]}.zip"'
        },
    )


class BatchEntry(BaseModel):
    item_code: str
    batch_sizes: list[float]


class BatchRunRequest(BaseModel):
    entries: list[BatchEntry]
    notes: Optional[str] = None


def _write_batching_xlsx(entries: list[BatchEntry]) -> None:
    """Write batching_recipes.xlsx that RecipeBatchPrinter.py reads."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item Code", "Batch Size"])
    for entry in entries:
        if not entry.batch_sizes:
            continue
        # Convert [1500, 1500, 1500] → "1500/1500/1500"
        batch_str = "/".join(str(int(s)) for s in entry.batch_sizes)
        ws.append([entry.item_code, batch_str])
    dest = RECIPE_AUTOMATION_DIR / "batching_recipes.xlsx"
    wb.save(dest)


@app.post("/api/runs/from-batches", response_model=RunStatus, status_code=202)
async def run_from_batches(body: BatchRunRequest, background_tasks: BackgroundTasks) -> RunStatus:
    """Accept pre-calculated batch sizes, generate batching_recipes.xlsx, run pipeline."""
    _write_batching_xlsx(body.entries)
    run_id = str(uuid.uuid4())
    supabase.table("runs").insert(
        {"id": run_id, "status": "queued", "notes": body.notes}
    ).execute()
    background_tasks.add_task(run_pipeline_recipes_only, run_id)
    return RunStatus(run_id=run_id, status="queued")


async def run_pipeline_recipes_only(run_id: str) -> None:
    """Run only RecipeBatchPrinter.py (batching_recipes.xlsx already written)."""
    temp_dir = RECIPE_AUTOMATION_DIR / "temp_print"
    try:
        supabase.table("runs").update({"status": "running"}).eq("id", run_id).execute()
        script_path = RECIPE_AUTOMATION_DIR / "RecipeBatchPrinter.py"
        if script_path.exists():
            await _run_script(script_path, RECIPE_AUTOMATION_DIR)

        if temp_dir.exists():
            recipe_file_records: list[dict] = []
            ingredient_records: list[dict] = []
            for xlsx_path in sorted(temp_dir.glob("*.xlsx")):
                filename = xlsx_path.name
                meta = _classify_file(filename)
                if meta is None:
                    continue
                storage_path = f"{run_id}/{filename}"
                with open(xlsx_path, "rb") as fh:
                    supabase.storage.from_(STORAGE_BUCKET).upload(
                        storage_path, fh.read(),
                        {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         "upsert": "true"},
                    )
                recipe_file_records.append({"run_id": run_id, "file_name": filename,
                                             "storage_path": storage_path, **meta})
                for row in _parse_recipe_file(xlsx_path):
                    ingredient_records.append({"run_id": run_id, "file_name": filename,
                                                "item_code": meta["item_code"],
                                                "file_type": meta["file_type"],
                                                "batch_number": meta["batch_number"],
                                                "batch_size_kg": meta["batch_size_kg"], **row})
            if recipe_file_records:
                supabase.table("recipe_files").insert(recipe_file_records).execute()
            if ingredient_records:
                supabase.table("recipe_ingredients").insert(ingredient_records).execute()
            for f in temp_dir.glob("*.xlsx"):
                f.unlink(missing_ok=True)

        supabase.table("runs").update({"status": "done"}).eq("id", run_id).execute()
    except Exception as exc:
        supabase.table("runs").update(
            {"status": "error", "error_message": str(exc)}
        ).eq("id", run_id).execute()
        raise


@app.post("/api/runs/upload", response_model=RunStatus, status_code=202)
async def upload_and_run(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    notes: Optional[str] = None,
) -> RunStatus:
    """Accept a production xlsx upload, save it, then kick off the pipeline."""
    dest = RECIPE_AUTOMATION_DIR / "production.xlsx"
    dest.write_bytes(await file.read())

    run_id = str(uuid.uuid4())
    supabase.table("runs").insert(
        {"id": run_id, "status": "queued", "notes": notes}
    ).execute()
    background_tasks.add_task(run_pipeline, run_id, str(dest))
    return RunStatus(run_id=run_id, status="queued")


# ─── Keywords that identify Line 7-8 files (case-insensitive substring match) ─
_LINE78_KEYWORDS = ["line 7", "line7", "7-8", "7_8", "l7_", "l78", "line78"]


@app.post("/api/webhook/email", status_code=202)
async def email_webhook(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    notes: Optional[str] = Form(None),
    x_webhook_secret: Optional[str] = Header(None),
) -> dict:
    """
    Called by Make.com (or similar) for each xlsx attachment from the
    production email.  Non-xlsx files and Line 7-8 sheets are skipped
    so only the main production plan triggers a pipeline run.

    Make.com HTTP module setup:
      URL    : POST https://<your-backend>/api/webhook/email
      Headers: x-webhook-secret: <WEBHOOK_SECRET>
      Body   : multipart/form-data
                 file  = {{attachment binary}}
                 notes = {{email subject}} – {{date}}
    """
    # ── Auth ──────────────────────────────────────────────────────────────────
    if WEBHOOK_SECRET and x_webhook_secret != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Invalid webhook secret")

    filename = file.filename or ""

    # ── Skip non-xlsx files (PDFs, images, Word docs, etc.) ───────────────────
    if not filename.lower().endswith(".xlsx"):
        return {"skipped": True, "reason": "not an xlsx file", "filename": filename}

    # ── Skip Line 7-8 sheets — user uploads these manually in the board ───────
    name_lower = filename.lower()
    if any(kw in name_lower for kw in _LINE78_KEYWORDS):
        return {"skipped": True, "reason": "line 7-8 file — upload via the board", "filename": filename}

    # ── Treat as production plan — same logic as /api/runs/upload ─────────────
    dest = RECIPE_AUTOMATION_DIR / "production.xlsx"
    dest.write_bytes(await file.read())

    run_id = str(uuid.uuid4())
    run_notes = notes or f"Auto-uploaded from email: {filename}"
    supabase.table("runs").insert(
        {"id": run_id, "status": "queued", "notes": run_notes}
    ).execute()
    background_tasks.add_task(run_pipeline, run_id, str(dest))

    return {"queued": True, "run_id": run_id, "filename": filename}


@app.get("/api/runs")
async def list_runs(limit: int = 10) -> list:
    """Return the most recent runs (newest first)."""
    result = (
        supabase.table("runs")
        .select("id, status, notes, created_at")
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    return result.data or []


@app.get("/api/runs/{run_id}/recipe-files")
async def get_recipe_files(run_id: str) -> list:
    """Return all recipe_files rows for a run (service-role key bypasses RLS)."""
    result = (
        supabase.table("recipe_files")
        .select("file_name, item_code, file_type, batch_number, batch_size_kg")
        .eq("run_id", run_id)
        .order("item_code")
        .order("batch_number")
        .execute()
    )
    return result.data or []


@app.get("/api/runs/{run_id}/ingredients")
async def get_ingredients(run_id: str, file_names: str = "") -> list:
    """Return recipe_ingredients rows for a run, optionally filtered by file_names (comma-separated)."""
    query = (
        supabase.table("recipe_ingredients")
        .select("file_name, mix, ingredient_code, description, percentage, new_batch_qty")
        .eq("run_id", run_id)
    )
    if file_names:
        names = [n.strip() for n in file_names.split(",") if n.strip()]
        if names:
            query = query.in_("file_name", names)
    result = query.execute()
    return result.data or []
